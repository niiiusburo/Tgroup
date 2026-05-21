#!/usr/bin/env python3
"""
NK2 (or NK) Regression Test Suite

Exercises every bug fix from the v0.32.6 → v0.32.13 feedback-stabilization work
against a live deployment. Runs in <2 minutes. Exits 0 if all pass, 1 if any
fail — so CI / pre-deploy gates can use it.

Usage:
    # Default: hit NK2 staging
    python3 scripts/nk2-regression.py

    # Test against NK production (after deploy)
    NK_BASE=https://nk.2checkin.com python3 scripts/nk2-regression.py

    # Use a different admin account
    NK_EMAIL=t@clinic.vn NK_PASSWORD=123123 python3 scripts/nk2-regression.py

What it checks (one test per session-fixed bug):
    [v0.32.6]  Hosoonline endpoint reachable + returns expected envelope
    [v0.32.7]  Excel: Phiếu khám column has SO codes (not service names)
    [v0.32.10] Excel: Revenue has receipt_number + per-SO totals columns
    [v0.32.11] Excel: Revenue 'Tên dịch vụ' resolves real names for old SOs
    [v0.32.11] Excel: Payments 'Số phiếu điều trị' shows SO codes only
    [v0.32.12] Excel: Payments 'Số phiếu điều trị' populated for ≥80% of rows
    [v0.32.12] API:  Sales-employees report accepts legacy GUID company filter
    [v0.32.13] API:  Customer resolver returns 200 for code, phone, UUID
    [v0.32.13] API:  Customer resolver returns 404 for unknown key
    [v0.32.13] API:  Customer resolver returns 400 for missing key
    [smoke]    All 8 Excel export types download as valid xlsx files
"""
from __future__ import annotations
import json, os, sys, tempfile, urllib.parse, urllib.request, urllib.error, re
from pathlib import Path

NK_BASE = os.environ.get("NK_BASE", "https://nk2.2checkin.com").rstrip("/")
EMAIL = os.environ.get("NK_EMAIL", "t@clinic.vn")
PASSWORD = os.environ.get("NK_PASSWORD", "123123")

# Known test fixtures on NK2 staging (and NK production)
KNOWN_CUSTOMER_CODE = "T056733"  # LÊ HUỲNH MỸ TRÂM
KNOWN_CUSTOMER_PHONE = "0939361880"
KNOWN_CUSTOMER_UUID = "61d6759f-f2e2-4443-8d5b-b3a7003ab7c5"
KNOWN_COMPANY_UUID = "c6b4b453-d260-46d4-4fd9-08db24f7ae8e"  # Tấm Dentist Quận 3 (legacy GUID format)

GREEN = "\033[32m"
RED = "\033[31m"
YELLOW = "\033[33m"
DIM = "\033[2m"
RESET = "\033[0m"

results: list[tuple[str, bool, str]] = []  # (name, passed, detail)


def http_post_json(path: str, payload: dict, token: str | None = None) -> tuple[int, dict | bytes]:
    body = json.dumps(payload).encode()
    headers = {"Content-Type": "application/json"}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    req = urllib.request.Request(f"{NK_BASE}{path}", data=body, headers=headers, method="POST")
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            raw = r.read()
            ctype = r.headers.get("Content-Type", "")
            if "application/json" in ctype:
                return r.status, json.loads(raw)
            return r.status, raw
    except urllib.error.HTTPError as e:
        raw = e.read()
        try: return e.code, json.loads(raw)
        except Exception: return e.code, raw


def http_get(path: str, token: str | None = None) -> tuple[int, dict | bytes]:
    headers = {}
    if token:
        headers["Authorization"] = f"Bearer {token}"
    req = urllib.request.Request(f"{NK_BASE}{path}", headers=headers)
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            raw = r.read()
            ctype = r.headers.get("Content-Type", "")
            if "application/json" in ctype:
                return r.status, json.loads(raw)
            return r.status, raw
    except urllib.error.HTTPError as e:
        raw = e.read()
        try: return e.code, json.loads(raw)
        except Exception: return e.code, raw


def login() -> str:
    code, body = http_post_json("/api/Auth/login", {"email": EMAIL, "password": PASSWORD})
    if code != 200 or not isinstance(body, dict):
        raise SystemExit(f"Login failed: HTTP {code} {body}")
    tok = body.get("token") or body.get("access_token")
    if not tok:
        raise SystemExit(f"Login returned no token: keys={list(body.keys())}")
    return tok


def export_download(token: str, export_type: str, filters: dict) -> Path:
    code, body = http_post_json(f"/api/Exports/{export_type}/download", {"filters": filters}, token)
    if code != 200 or isinstance(body, dict):
        raise RuntimeError(f"Export {export_type} HTTP {code}: {body if isinstance(body, dict) else 'binary'}")
    tmp = Path(tempfile.mkstemp(suffix=f"_{export_type}.xlsx")[1])
    tmp.write_bytes(body)
    return tmp


def load_xlsx_data_sheet(path: Path, header_search_rows: int = 6):
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet = "Data" if "Data" in wb.sheetnames else ("Sheet1" if "Sheet1" in wb.sheetnames else wb.sheetnames[0])
    ws = wb[sheet]
    rows = []
    for i, r in enumerate(ws.iter_rows(max_row=max(2000, header_search_rows), values_only=True)):
        rows.append(r)
    wb.close()
    # find header row
    header_idx = None
    for i, r in enumerate(rows[:header_search_rows]):
        non_empty = sum(1 for v in r if v is not None and str(v).strip())
        if non_empty >= 3:
            header_idx = i; break
    if header_idx is None:
        return [], []
    return rows[header_idx], rows[header_idx+1:]


def pass_(name: str, detail: str = ""):
    results.append((name, True, detail))
    print(f"  {GREEN}PASS{RESET}  {name}" + (f"  {DIM}{detail}{RESET}" if detail else ""))


def fail(name: str, detail: str):
    results.append((name, False, detail))
    print(f"  {RED}FAIL{RESET}  {name}  {detail}")


# ---------------------------- Tests ----------------------------
def test_hosoonline_endpoint(token: str):
    code, body = http_get(f"/api/ExternalCheckups/{KNOWN_CUSTOMER_CODE}", token)
    if code != 200 or not isinstance(body, dict):
        return fail("hosoonline_endpoint", f"HTTP {code}")
    if "checkups" not in body or "source" not in body:
        return fail("hosoonline_endpoint", f"missing keys: got {list(body.keys())}")
    pass_("hosoonline_endpoint", f"source={body['source']} checkups={len(body.get('checkups', []))}")


def test_customer_resolver(token: str):
    # ref
    code, body = http_get(f"/api/Partners/resolve?key={KNOWN_CUSTOMER_CODE}", token)
    if code != 200 or not isinstance(body, dict) or body.get("matchedBy") != "ref":
        return fail("resolver_ref", f"HTTP {code} body={body}")
    pass_("resolver_ref", f"{KNOWN_CUSTOMER_CODE} → {body['partner']['id']}")

    # phone
    code, body = http_get(f"/api/Partners/resolve?key={KNOWN_CUSTOMER_PHONE}", token)
    if code != 200 or not isinstance(body, dict) or body.get("matchedBy") != "phone":
        return fail("resolver_phone", f"HTTP {code} body={body}")
    pass_("resolver_phone", f"{KNOWN_CUSTOMER_PHONE} → {body['partner']['id']}")

    # uuid passthrough
    code, body = http_get(f"/api/Partners/resolve?key={KNOWN_CUSTOMER_UUID}", token)
    if code != 200 or not isinstance(body, dict) or body.get("matchedBy") != "uuid":
        return fail("resolver_uuid", f"HTTP {code} body={body}")
    pass_("resolver_uuid", "UUID passthrough OK")

    # 404
    code, body = http_get(f"/api/Partners/resolve?key=NOTAREALCODE_{os.getpid()}", token)
    if code != 404 or not isinstance(body, dict) or body.get("code") != "CUSTOMER_NOT_FOUND":
        return fail("resolver_404", f"HTTP {code} body={body}")
    pass_("resolver_404", "unknown key → 404 CUSTOMER_NOT_FOUND")

    # 400 missing
    code, body = http_get("/api/Partners/resolve", token)
    if code != 400 or not isinstance(body, dict) or body.get("code") != "CUSTOMER_LOOKUP_KEY_REQUIRED":
        return fail("resolver_400", f"HTTP {code} body={body}")
    pass_("resolver_400", "missing key → 400")


def test_sales_employees_company_filter(token: str):
    code, body = http_post_json("/api/Exports/report-sales-employees/download", {
        "filters": {"dateFrom": "2026-04-16", "dateTo": "2026-05-16", "companyId": KNOWN_COMPANY_UUID}
    }, token)
    if code != 200:
        return fail("sales_employees_company_filter", f"HTTP {code} body={body}")
    pass_("sales_employees_company_filter", f"company filter accepted, returned {len(body) if isinstance(body, bytes) else '?'} bytes")


def test_revenue_export_column_F(token: str):
    try:
        f = export_download(token, "revenue-flat", {"dateFrom": "2026-04-16", "dateTo": "2026-05-16", "companyId": "all"})
    except Exception as e:
        return fail("revenue_F_no_mirror", f"download failed: {e}")
    headers, data_rows = load_xlsx_data_sheet(f)
    if not headers:
        return fail("revenue_F_no_mirror", "no data sheet header")
    same = 0
    for r in data_rows[:1000]:
        e = str(r[4] or "") if len(r) > 4 else ""
        fcol = str(r[5] or "") if len(r) > 5 else ""
        if e.startswith("SO") and e == fcol:
            same += 1
    if same > 0:
        return fail("revenue_F_no_mirror", f"{same}/1000 rows have E==F (SO code leaking into name col)")
    pass_("revenue_F_no_mirror", f"0/1000 mirror rows")


def test_revenue_has_v0_32_10_columns(token: str):
    try:
        f = export_download(token, "revenue-flat", {"dateFrom": "2026-04-16", "dateTo": "2026-05-16", "companyId": "all"})
    except Exception as e:
        return fail("revenue_has_v32_10_cols", f"download failed: {e}")
    headers, _ = load_xlsx_data_sheet(f)
    header_text = " | ".join(str(h or "") for h in headers).lower()
    missing = []
    for col in ["biên lai", "tổng tiền phiếu", "còn lại phiếu"]:
        if col not in header_text: missing.append(col)
    if missing:
        return fail("revenue_has_v32_10_cols", f"missing columns: {missing}")
    pass_("revenue_has_v32_10_cols", "biên lai + tổng/còn lại phiếu columns present")


def test_payments_col_format(token: str):
    try:
        f = export_download(token, "payments", {"dateFrom": "2026-04-16", "dateTo": "2026-05-16", "companyId": "all"})
    except Exception as e:
        return fail("payments_col_format", f"download failed: {e}")
    headers, data_rows = load_xlsx_data_sheet(f)
    # locate "Số phiếu điều trị" column
    col_idx = None
    for i, h in enumerate(headers):
        if h and "số phiếu điều trị" in str(h).lower():
            col_idx = i; break
    if col_idx is None:
        return fail("payments_col_format", "column 'Số phiếu điều trị' not found")
    bad = 0; populated = 0; total = 0
    for r in data_rows[:1000]:
        if col_idx >= len(r): continue
        total += 1
        v = str(r[col_idx] or "").strip()
        if v:
            populated += 1
            for part in v.split(","):
                if part.strip() and not part.strip().startswith("SO"):
                    bad += 1; break
    if bad > 0:
        return fail("payments_col_format", f"{bad}/{total} rows have non-SO values in SO column")
    pass_("payments_col_format", f"{populated}/{total} populated, 0 bad-format")


def test_payments_col_populated(token: str, min_pct: float = 80.0):
    # Same download but check populate rate
    try:
        f = export_download(token, "payments", {"dateFrom": "2026-04-16", "dateTo": "2026-05-16", "companyId": "all"})
    except Exception as e:
        return fail(f"payments_col_populated", f"download failed: {e}")
    headers, data_rows = load_xlsx_data_sheet(f)
    col_idx = next((i for i, h in enumerate(headers) if h and "số phiếu điều trị" in str(h).lower()), None)
    if col_idx is None: return fail("payments_col_populated", "column not found")
    populated = total = 0
    for r in data_rows[:1000]:
        if col_idx >= len(r): continue
        total += 1
        if r[col_idx] and str(r[col_idx]).strip(): populated += 1
    if total == 0: return fail("payments_col_populated", "no data rows")
    pct = populated * 100 / total
    if pct < min_pct:
        return fail("payments_col_populated", f"only {pct:.1f}% populated (need ≥{min_pct}%)")
    pass_("payments_col_populated", f"{pct:.1f}% of rows populated")


def test_payments_phieukham_is_so_code(token: str):
    """v0.32.7: 'Phiếu khám' / SO code column should hold SO codes, not service names."""
    try:
        f = export_download(token, "revenue-flat", {"dateFrom": "2026-04-16", "dateTo": "2026-05-16", "companyId": "all"})
    except Exception as e:
        return fail("phieukham_is_so_code", f"download failed: {e}")
    _, data_rows = load_xlsx_data_sheet(f)
    bad = 0
    for r in data_rows[:1000]:
        v = str(r[4] or "").strip() if len(r) > 4 else ""  # col E
        if v and not (v.startswith("SO") or v == ""):
            bad += 1
    if bad > 0:
        return fail("phieukham_is_so_code", f"{bad}/1000 rows in col E are not SO codes")
    pass_("phieukham_is_so_code", "all populated col E rows start with 'SO'")


def test_all_exports_valid_xlsx(token: str):
    types = ["service-catalog", "customers", "appointments", "services", "payments",
             "report-sales-employees", "revenue-flat", "deposit-flat"]
    bad = []
    for t in types:
        try:
            f = export_download(token, t, {"dateFrom": "2026-04-16", "dateTo": "2026-05-16", "companyId": "all"})
            # quick magic-bytes check
            with open(f, "rb") as fh:
                sig = fh.read(4)
            if sig[:2] != b"PK":
                bad.append(f"{t} (not zip)")
        except Exception as e:
            bad.append(f"{t} ({e})")
    if bad:
        return fail("all_exports_valid", f"failed: {bad}")
    pass_("all_exports_valid", f"all {len(types)} exports return valid xlsx")


# ---------------------------- Main ----------------------------
def main():
    print(f"\nNK2 Regression Suite  →  {NK_BASE}\n")

    # version
    try:
        code, body = http_get("/version.json")
        if code == 200 and isinstance(body, dict):
            print(f"  Target version: {body.get('version', '?')}  built {body.get('buildTime', '?')}")
    except Exception:
        pass

    print(f"  Logging in as {EMAIL}...")
    token = login()
    print(f"  Token acquired.\n")

    # Run all tests
    print(f"{DIM}--- API tests ---{RESET}")
    test_hosoonline_endpoint(token)
    test_customer_resolver(token)
    test_sales_employees_company_filter(token)

    print(f"\n{DIM}--- Excel export tests ---{RESET}")
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        print(f"  {YELLOW}SKIP{RESET}  Excel tests — openpyxl not installed (pip install openpyxl)")
    else:
        test_phieukham_is_so_code = test_payments_phieukham_is_so_code  # alias for clarity
        test_phieukham_is_so_code(token)
        test_revenue_has_v0_32_10_columns(token)
        test_revenue_export_column_F(token)
        test_payments_col_format(token)
        test_payments_col_populated(token)
        test_all_exports_valid_xlsx(token)

    # Summary
    print()
    passed = sum(1 for _, ok, _ in results if ok)
    failed = sum(1 for _, ok, _ in results if not ok)
    total = passed + failed
    if failed == 0:
        print(f"{GREEN}━━━ {passed}/{total} PASSED ━━━{RESET}")
        sys.exit(0)
    else:
        print(f"{RED}━━━ {failed}/{total} FAILED ━━━{RESET}")
        for n, ok, d in results:
            if not ok: print(f"  {RED}✗{RESET} {n}  {d}")
        sys.exit(1)


if __name__ == "__main__":
    main()
