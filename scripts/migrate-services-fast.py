#!/usr/bin/env python3.13
"""
Fast service migration using psycopg2 (batch updates)
Excel → dbo.saleorders + dbo.saleorderlines
"""
import openpyxl
import psycopg2
from psycopg2.extras import execute_batch
from collections import defaultdict

DB = {
    'host': '127.0.0.1',
    'port': 5433,
    'user': 'postgres',
    'password': 'postgres',
    'dbname': 'tdental_demo',
}

def get_conn():
    return psycopg2.connect(**DB)

print("Loading Excel...")
wb = openpyxl.load_workbook('/Users/thuanle/Downloads/Phiếu điều trị.xlsx', read_only=True, data_only=True)
so_data = defaultdict(dict)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows(min_row=2, values_only=True):
        so = str(row[0]).strip() if row[0] else ''
        val = str(row[1]).strip() if row[1] else ''
        if so and val and sheet_name not in so_data[so]:
            so_data[so][sheet_name] = val
wb.close()
print(f"  Unique SOs: {len(so_data)}")

conn = get_conn()
cur = conn.cursor()

print("Loading employees...")
cur.execute("""
    SELECT id, ref, UPPER(TRIM(name))
    FROM dbo.partners
    WHERE employee = true AND COALESCE(TRIM(name), '') <> ''
""")
emp_rows = cur.fetchall()
emp_candidates = defaultdict(list)
for eid, ref, name in emp_rows:
    emp_candidates[name].append((eid, ref))

ambiguous_employees = {name: rows for name, rows in emp_candidates.items() if len(rows) > 1}
if ambiguous_employees:
    print("\nERROR: Duplicate employee names found. Refine the mapping before running this migration.")
    for name, rows in sorted(ambiguous_employees.items())[:25]:
        refs = ", ".join(ref or "no-ref" for _, ref in rows)
        print(f"  {name}: {refs}")
    print(f"  Total ambiguous names: {len(ambiguous_employees)}")
    raise SystemExit(1)

emp_map = {name: rows[0][0] for name, rows in emp_candidates.items()}
print(f"  {len(emp_map)} employees")

print("Loading sources...")
cur.execute("SELECT id, UPPER(TRIM(name)) FROM dbo.customersources")
source_map = {name: sid for sid, name in cur.fetchall()}
print(f"  {len(source_map)} sources")

print("Loading saleorders...")
cur.execute("SELECT id, name FROM dbo.saleorders")
so_map = {name: soid for soid, name in cur.fetchall()}
print(f"  {len(so_map)} saleorders")

# Build update lists. Database contract:
# assistantid = Phu ta, dentalaideid = Tro ly bac si.
order_updates = []   # (doctorid, assistantid, dentalaideid, sourceid, soid)
line_updates = []    # (assistantid, soid)

skipped = 0
for sonum, data in so_data.items():
    if sonum not in so_map:
        skipped += 1
        continue
    so_id = so_map[sonum]
    doc_id = emp_map.get(data.get('Bác sĩ', '').strip().upper())
    assistant_id = emp_map.get(data.get('Phụ tá', '').strip().upper())
    dental_aide_id = emp_map.get(data.get('Trợ lý bác sĩ', '').strip().upper())
    src_id = source_map.get(data.get('Nguồn', '').strip().upper())
    
    if any([doc_id, assistant_id, dental_aide_id, src_id]):
        order_updates.append((doc_id, assistant_id, dental_aide_id, src_id, so_id))
    if assistant_id:
        line_updates.append((assistant_id, so_id))

print(f"  Matched: {len(order_updates)} orders, {len(line_updates)} line batches")
print(f"  Skipped (no DB match): {skipped}")

print("\nUpdating saleorders...")
BATCH = 1000
updated_orders = 0
for i in range(0, len(order_updates), BATCH):
    batch = order_updates[i:i+BATCH]
    execute_batch(cur, """
        UPDATE dbo.saleorders
        SET doctorid = COALESCE(%s, doctorid),
            assistantid = COALESCE(%s, assistantid),
            dentalaideid = COALESCE(%s, dentalaideid),
            sourceid = COALESCE(%s, sourceid)
        WHERE id = %s
    """, batch, page_size=BATCH)
    updated_orders += len(batch)
    conn.commit()
    if (i // BATCH + 1) % 10 == 0:
        print(f"  Batch {i//BATCH + 1}/{(len(order_updates)+BATCH-1)//BATCH}")

print(f"  Updated {updated_orders} saleorders")

print("\nUpdating saleorderlines...")
updated_lines = 0
for i in range(0, len(line_updates), BATCH):
    batch = line_updates[i:i+BATCH]
    execute_batch(cur, """
        UPDATE dbo.saleorderlines
        SET assistantid = %s
        WHERE orderid = %s
    """, batch, page_size=BATCH)
    updated_lines += len(batch)
    conn.commit()
    if (i // BATCH + 1) % 10 == 0:
        print(f"  Batch {i//BATCH + 1}/{(len(line_updates)+BATCH-1)//BATCH}")

print(f"  Updated {updated_lines} saleorderline groups")

# Verification
print("\n--- Verification ---")
cur.execute("SELECT COUNT(*) FROM dbo.saleorders WHERE doctorid IS NOT NULL")
print(f"  Orders with doctor: {cur.fetchone()[0]}")
cur.execute("SELECT COUNT(*) FROM dbo.saleorders WHERE dentalaideid IS NOT NULL")
print(f"  Orders with doctor assistant: {cur.fetchone()[0]}")
cur.execute("SELECT COUNT(*) FROM dbo.saleorders WHERE assistantid IS NOT NULL")
print(f"  Orders with assistant: {cur.fetchone()[0]}")
cur.execute("SELECT COUNT(*) FROM dbo.saleorders WHERE sourceid IS NOT NULL")
print(f"  Orders with source: {cur.fetchone()[0]}")
cur.execute("SELECT COUNT(*) FROM dbo.saleorderlines WHERE assistantid IS NOT NULL")
print(f"  Lines with assistant: {cur.fetchone()[0]}")

cur.close()
conn.close()
print("\n✅ Service migration complete")
