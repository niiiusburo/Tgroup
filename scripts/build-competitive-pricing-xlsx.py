#!/usr/bin/env python3.12
"""Build TMV Tấm competitive pricing workbook (anchor SKUs vs 20 VN competitors)."""
import json
import re
import statistics
import urllib.request
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

OUT = Path(__file__).resolve().parents[1] / "docs/reports/TMV-Tam-competitive-pricing-vn-2026-06-09.xlsx"
PRICING_URL = "https://tmv.2checkin.com/bang-gia/data/pricing.json"

ANCHORS = [
    ("filler_han_1cc", "Thẩm mỹ", "Filler Hàn 1cc", "1cc filler cơ bản (Hàn)"),
    ("filler_premium_1cc", "Thẩm mỹ", "Filler cao cấp 1cc", "1cc filler premium"),
    ("juvederm_1cc", "Thẩm mỹ", "Juvederm 1cc", "Juvederm / filler cao cấp"),
    ("combo_moi_2cc", "Thẩm mỹ", "Combo môi filler 2cc", "Gói môi 2cc"),
    ("botox_han_ham", "Thẩm mỹ", "Botox Hàn gọn hàm", "~50 unit"),
    ("botox_us_ham", "Thẩm mỹ", "Botox Mỹ gọn hàm", "Botox Mỹ thon hàm"),
    ("nose_entry", "Thẩm mỹ", "Nâng mũi entry (Hàn/basic)", "Mức entry nâng mũi"),
    ("nose_4d", "Thẩm mỹ", "Nâng mũi cấu trúc 4D", "Softline / Nanoform tier"),
    ("eyelid_basic", "Thẩm mỹ", "Cắt mí basic", "Cắt mí / mí Hàn"),
    ("breast_nano", "Thẩm mỹ", "Nâng ngực Nano Chip", "Nâng ngực nano chip"),
    ("implant_kr", "Nha khoa", "Implant Hàn (trụ+abutment)", "Gói implant HQ"),
    ("implant_ch", "Nha khoa", "Implant Thụy Sỹ", "Gói implant Thụy Sỹ"),
    ("braces_metal", "Nha khoa", "Niềng kim loại (cơ bản)", "Mắc cài kim loại"),
    ("braces_ceramic", "Nha khoa", "Niềng sứ (cơ bản)", "Mắc cài sứ"),
    ("porcelain_titan", "Nha khoa", "Răng sứ Titan", "1 răng titan"),
    ("porcelain_zirconia", "Nha khoa", "Răng sứ Zirconia", "1 răng zirconia"),
]

COMPETITORS = [
    "TMV Tấm",
    "Gangwhoo",
    "Kangnam",
    "Thu Cúc",
    "JW Korea",
    "Linh Anh",
    "Seoul Center",
    "Lavender",
    "Saigon Cosmetic",
    "Doctor Laser",
    "Shinbi",
    "Xuân Hướng",
    "Tâm Anh",
    "Worldwide",
    "Nha Khoa Kim",
    "Paris Dental",
    "Flora",
    "Viet Smile",
    "DK Dental",
    "Westway",
]

# price_vnd, source_type (official|secondary), comparable_label, notes
COMP_DATA = {
    "Gangwhoo": {
        "nose_entry": (22_000_000, "official", "Nâng mũi Hàn Quốc (list)", "Promo 14M tháng 6/2026"),
        "nose_4d": (44_000_000, "official", "Sline/Lline 4D (list)", "Promo 46.2M→35M khác mục"),
        "eyelid_basic": (11_000_000, "official", "Mí Hàn Quốc (list)", "Promo 8M"),
    },
    "Kangnam": {
        "nose_4d": (35_200_000, "official", "Nâng mũi Nanoform Idol", "Kangnam price list"),
        "eyelid_basic": (12_000_000, "official", "Cắt mí 6D (Easy)", ""),
        "breast_nano": (67_000_000, "official", "Nâng ngực 6D Nano chip", ""),
    },
    "Thu Cúc": {
        "filler_han_1cc": (3_000_000, "official", "Filler TC 1cc", ""),
        "juvederm_1cc": (12_000_000, "official", "Filler TC Juvederm", ""),
        "botox_han_ham": (8_500_000, "official", "Botox gọn hàm (trọn gói)", "20–60 UI"),
        "nose_entry": (18_000_000, "official", "Nâng mũi TC 2024", "benhvienthammythucuc.vn"),
        "nose_4d": (20_000_000, "official", "Nâng mũi cấu trúc", "Mục cấu trúc bảng giá"),
    },
    "JW Korea": {
        "eyelid_basic": (17_000_000, "official", "Cắt mắt 2 mí", "Giá tham khảo JW"),
    },
    "Nha Khoa Kim": {
        "implant_kr": (21_000_000, "official", "Implant HQ + abutment", "Biotem/Dio/Megagen"),
        "implant_ch": (30_000_000, "official", "Implant Thụy Sỹ", "Straumann/Swiss"),
        "braces_metal": (40_000_000, "official", "Niềng kim loại đơn giản", ""),
        "braces_ceramic": (55_000_000, "official", "Niềng sứ đơn giản", ""),
        "porcelain_titan": (3_000_000, "official", "Răng sứ kim loại Titan", ""),
        "porcelain_zirconia": (6_000_000, "official", "Zirconia Argon Multilayer", ""),
    },
    "Paris Dental": {
        "implant_kr": (12_000_000, "official", "Trụ Implant Dio HQ", "Trụ only; update 2025"),
        "implant_ch": (30_000_000, "official", "Trụ Implant SIC/Straumann", ""),
        "braces_metal": (30_000_000, "official", "Niềng KL 3M cơ bản", ""),
        "braces_ceramic": (48_000_000, "official", "Niềng sứ AI DESIGN cơ bản", ""),
        "porcelain_titan": (3_000_000, "official", "Răng sứ Venus", ""),
        "porcelain_zirconia": (6_000_000, "official", "Emax Zic / Cercon", ""),
    },
    "Flora": {
        "implant_kr": (14_500_000, "secondary", "Implant Hàn (ước tính)", "Cần xác minh bảng giá Flora"),
        "braces_metal": (35_000_000, "secondary", "Niềng kim loại (ước tính)", ""),
        "porcelain_zirconia": (5_500_000, "secondary", "Zirconia (ước tính)", ""),
    },
    "Viet Smile": {
        "implant_kr": (13_000_000, "secondary", "Implant Hàn (ước tính)", ""),
        "braces_metal": (32_000_000, "secondary", "Niềng kim loại (ước tính)", ""),
    },
    "DK Dental": {
        "implant_kr": (15_000_000, "secondary", "Implant Hàn (ước tính)", ""),
        "porcelain_zirconia": (5_000_000, "secondary", "Zirconia (ước tính)", ""),
    },
    "Westway": {
        "implant_ch": (28_000_000, "secondary", "Implant Thụy Sỹ (ước tính)", ""),
        "braces_ceramic": (50_000_000, "secondary", "Niềng sứ (ước tính)", ""),
    },
}

SOURCE_URLS = {
    "TMV Tấm": "https://tmv.2checkin.com/bang-gia/data/pricing.json",
    "Gangwhoo": "https://benhvienthammygangwhoo.vn/bang-gia/",
    "Kangnam": "https://kangnamaesthetichospital.com/tin-tuc/bang-gia-phau-thuat-tham-my-tai-kangnam/",
    "Thu Cúc": "https://benhvienthammythucuc.vn/gia-tiem-filler/",
    "JW Korea": "https://benhvienjw.vn/bang-gia-tham-my",
    "Linh Anh": "https://thammylinhanh.vn/",
    "Seoul Center": "https://seoulcenter.vn/",
    "Lavender": "https://lavender.com.vn/",
    "Saigon Cosmetic": "https://saigoncosmetic.com/",
    "Doctor Laser": "https://doctorlaser.vn/",
    "Shinbi": "https://shinbi.vn/",
    "Xuân Hướng": "https://thammyxuanhuong.com/",
    "Tâm Anh": "https://benhvientamanh.vn/",
    "Worldwide": "https://worldwidehospital.vn/",
    "Nha Khoa Kim": "https://nhakhoakim.com/bang-gia-nha-khoa-nieng-rang-rang-su-implant.html",
    "Paris Dental": "https://nhakhoaparis.vn/hoan-my-bang-gia-dich-vu-nha-khoa.html",
    "Flora": "https://floradental.com.vn/",
    "Viet Smile": "https://vietsmile.vn/",
    "DK Dental": "https://dkdental.com.vn/",
    "Westway": "https://westwaydental.com.vn/",
}


def parse_vnd(s):
    if not s:
        return None
    digits = re.sub(r"[^\d]", "", str(s))
    return int(digits) if digits else None


def load_tmv_prices():
    raw = urllib.request.urlopen(PRICING_URL, timeout=30).read()
    pricing = json.loads(raw)
    tmv_anchor = {}
    catalog = []

    def pick(cat_id, name_match=None, exact_name=None):
        cat = next((c for c in pricing["categories"] if c["id"] == cat_id), None)
        if not cat:
            return None
        for item in cat["items"]:
            name = item.get("name", "")
            if exact_name and name == exact_name:
                return parse_vnd(item.get("price"))
            if name_match and name_match.lower() in name.lower():
                return parse_vnd(item.get("price"))
        return None

    tmv_anchor["filler_han_1cc"] = pick("filler-basic", exact_name="1cc Filler Hàn")
    tmv_anchor["filler_premium_1cc"] = pick("filler-premium", exact_name="1cc")
    tmv_anchor["juvederm_1cc"] = pick("filler-premium", exact_name="1cc Filler Juverderm")
    tmv_anchor["combo_moi_2cc"] = pick("filler-basic", exact_name="COMBO Môi 2cc")
    tmv_anchor["botox_han_ham"] = pick("slimming", exact_name="Thon gọn hàm Botox Hàn 50 unit")
    tmv_anchor["botox_us_ham"] = pick("slimming", exact_name="Thon gọn hàm Botox Mỹ")
    tmv_anchor["nose_entry"] = pick("nose", exact_name="Nâng mũi Mini Line sụn Hàn thường")
    tmv_anchor["nose_4d"] = pick("nose", exact_name="Nâng mũi Softline 4D")
    tmv_anchor["eyelid_basic"] = pick("eyes", exact_name="Cắt mí Mini Deep")
    tmv_anchor["breast_nano"] = pick("breast", exact_name="Nâng ngực Nano Chip")
    tmv_anchor["implant_kr"] = pick("cay-ghep-implant", exact_name="Trụ Implant Hàn")
    tmv_anchor["implant_ch"] = pick("cay-ghep-implant", exact_name="Trụ Implant Thụy Sỹ")
    tmv_anchor["braces_metal"] = pick("mac-cai-canh-cam-tu-ong", exact_name="Mắc cài cánh cam tự đóng cấp độ 1")
    tmv_anchor["braces_ceramic"] = pick("mac-cai-canh-cam-tu-ong", exact_name="Mắc cài sứ tiêu chuẩn cấp độ 1")
    tmv_anchor["porcelain_titan"] = pick("rang-su-1-rang", exact_name="RS Titan")
    tmv_anchor["porcelain_zirconia"] = pick("rang-su-1-rang", exact_name="RS Ful Zirconia")

    for cat in pricing["categories"]:
        for item in cat["items"]:
            catalog.append(
                {
                    "category_id": cat["id"],
                    "category": cat["name"],
                    "name": item.get("name", ""),
                    "price_vnd": parse_vnd(item.get("price")),
                    "price_display": item.get("price", ""),
                    "bonus": item.get("bonus", ""),
                }
            )
    return tmv_anchor, catalog


def style_header(cell, fill="1F4E79"):
    cell.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def build():
    tmv_prices, catalog = load_tmv_prices()
    wb = Workbook()

    # --- Comparison sheet ---
    ws = wb.active
    ws.title = "So sánh anchor SKU"
    headers = [
        "Mã SKU",
        "Nhóm",
        "Dịch vụ TMV (anchor)",
        "Ghi chú so sánh",
        "TMV Tấm (VNĐ)",
        "Min thị trường",
        "Median thị trường",
        "Max thị trường",
        "TMV vs Median",
        "Vị thế TMV",
        "Số đối thủ có giá",
    ] + COMPETITORS[1:]

    for col, h in enumerate(headers, 1):
        c = ws.cell(1, col, h)
        style_header(c)

    comp_start_col = len(headers) - len(COMPETITORS) + 1
    tmv_col = 5
    min_col = 6
    med_col = 7
    max_col = 8
    vs_col = 9
    pos_col = 10
    count_col = 11

    for r, (sku, group, label, note) in enumerate(ANCHORS, start=2):
        ws.cell(r, 1, sku)
        ws.cell(r, 2, group)
        ws.cell(r, 3, label)
        ws.cell(r, 4, note)
        tmv_val = tmv_prices.get(sku)
        ws.cell(r, tmv_col, tmv_val)
        ws.cell(r, tmv_col).number_format = "#,##0"

        comp_prices = []
        for i, comp in enumerate(COMPETITORS[1:], start=0):
            col = comp_start_col + i
            entry = COMP_DATA.get(comp, {}).get(sku)
            if entry:
                price, src, _, _ = entry
                comp_prices.append(price)
                cell = ws.cell(r, col, price)
                cell.number_format = "#,##0"
                if src == "secondary":
                    cell.font = Font(color="C65911", name="Arial")
            else:
                ws.cell(r, col, None)

        if comp_prices:
            mn = min(comp_prices)
            mx = max(comp_prices)
            med = statistics.median(comp_prices)
            ws.cell(r, min_col, mn).number_format = "#,##0"
            ws.cell(r, med_col, med).number_format = "#,##0"
            ws.cell(r, max_col, mx).number_format = "#,##0"
            ws.cell(r, count_col, len(comp_prices))
            if tmv_val and med:
                ws.cell(r, vs_col, tmv_val / med - 1).number_format = "0.0%"
                if tmv_val < mn:
                    pos = "Dưới thị trường"
                elif tmv_val <= med:
                    pos = "Ở/bên dưới median"
                elif tmv_val <= mx:
                    pos = "Trên median"
                else:
                    pos = "Cao nhất"
                ws.cell(r, pos_col, pos)
            else:
                ws.cell(r, pos_col, "N/A")
        else:
            ws.cell(r, pos_col, "N/A")

        last_comp_col = comp_start_col + len(COMPETITORS) - 2
        for col in range(1, last_comp_col + 1):
            ws.cell(r, col).border = thin_border()
            ws.cell(r, col).alignment = Alignment(vertical="center", wrap_text=True)

    ws.freeze_panes = "F2"
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 14 if col > 10 else 18

    # --- Sources sheet ---
    src_ws = wb.create_sheet("Nguồn & ghi chú")
    src_headers = ["Đối thủ", "URL", "Ngày tham chiếu", "SKU", "Giá (VNĐ)", "Loại nguồn", "Dịch vụ tương đương", "Ghi chú"]
    for col, h in enumerate(src_headers, 1):
        c = src_ws.cell(1, col, h)
        style_header(c, fill="375623")

    row = 2
    src_ws.cell(row, 1, "TMV Tấm")
    src_ws.cell(row, 2, SOURCE_URLS["TMV Tấm"])
    src_ws.cell(row, 3, str(date.today()))
    src_ws.cell(row, 4, "ALL")
    src_ws.cell(row, 6, "official")
    src_ws.cell(row, 7, "Live pricing.json — 16 categories, 221 items")
    row += 1

    for comp in COMPETITORS[1:]:
        src_ws.cell(row, 1, comp)
        src_ws.cell(row, 2, SOURCE_URLS.get(comp, ""))
        src_ws.cell(row, 3, "2026-06-09")
        if comp not in COMP_DATA:
            src_ws.cell(row, 6, "no_public_price")
            src_ws.cell(row, 7, "Không có bảng giá công khai cho anchor SKU — cần báo giá")
            row += 1
            continue
        for sku, entry in COMP_DATA[comp].items():
            price, src, label, note = entry
            anchor_label = next(a[2] for a in ANCHORS if a[0] == sku)
            src_ws.cell(row, 4, sku)
            src_ws.cell(row, 5, price)
            src_ws.cell(row, 5).number_format = "#,##0"
            src_ws.cell(row, 6, src)
            src_ws.cell(row, 7, label or anchor_label)
            src_ws.cell(row, 8, note)
            row += 1

    # --- Summary sheet ---
    sum_ws = wb.create_sheet("Tóm tắt")
    sum_ws["A1"] = "TMV Tấm — So sánh giá với 20 đối thủ VN (anchor SKU)"
    sum_ws["A1"].font = Font(bold=True, size=14, name="Arial")
    sum_ws["A3"] = f"Ngày: {date.today().isoformat()}"
    sum_ws["A4"] = "Phương pháp: 16 anchor SKU (không phải full 221×20) vì đối thủ hiếm khi publish cùng tên dòng."
    sum_ws["A5"] = "Ô trống = không có giá công khai tương đương. Ô cam = giá secondary/ước tính — cần xác minh."
    bullets = [
        "",
        "Điểm nổi bật (có số từ nguồn official):",
        "• Filler Hàn 1cc: TMV 4.8M — Thu Cúc TC 3M (rẻ hơn); TMV cao hơn entry Thu Cúc ~60%",
        "• Botox gọn hàm Hàn: TMV 3.6M — Thu Cúc ~8.5M → TMV rẻ hơn đáng kể",
        "• Nâng mũi entry: TMV 15M — Gangwhoo list 22M (promo 14M); Thu Cúc TC 18M → TMV cạnh tranh",
        "• Cắt mí basic: TMV ~6M — Gangwhoo 11M, Kangnam 12M, JW 17M → TMV thấp hơn thị trường",
        "• Nâng ngực Nano: TMV 152M — Kangnam 67M (khác gói/chip) → không apples-to-apples",
        "• Implant Hàn: TMV 16.5M (trụ) — Kim 21M (trụ+abutment); Paris 12M (trụ) → so sánh theo scope",
        "• Niềng sứ cơ bản: TMV 30M — Kim 55M, Paris 48M → TMV thấp hơn",
        "• Zirconia 1 răng: TMV 3.69M — Kim 6M, Paris 6M → TMV thấp hơn ~40%",
        "",
        "Hạn chế:",
        "• Nhiều TMV đối thủ (Linh Anh, Seoul Center, Lavender, …) chỉ báo giá qua tư vấn",
        "• Promo/khuyến mãi làm lệch giá list — sheet dùng list price trừ khi ghi chú promo",
        "• Gói dịch vụ khác tên (Mini Line vs TC 2024 vs Nanoform) — chỉ so sánh tier gần nhất",
    ]
    for i, line in enumerate(bullets, start=7):
        sum_ws.cell(i, 1, line)
    sum_ws.column_dimensions["A"].width = 110

    # --- TMV catalog ---
    cat_ws = wb.create_sheet("TMV full catalog")
    cat_headers = ["category_id", "category", "item", "price_vnd", "price_display", "bonus"]
    for col, h in enumerate(cat_headers, 1):
        style_header(cat_ws.cell(1, col, h), fill="5B5B5B")
    for r, item in enumerate(catalog, start=2):
        cat_ws.cell(r, 1, item["category_id"])
        cat_ws.cell(r, 2, item["category"])
        cat_ws.cell(r, 3, item["name"])
        cat_ws.cell(r, 4, item["price_vnd"])
        cat_ws.cell(r, 4).number_format = "#,##0"
        cat_ws.cell(r, 5, item["price_display"])
        cat_ws.cell(r, 6, item["bonus"])

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Catalog items: {len(catalog)}")
    print(f"Anchors: {len(ANCHORS)} | Competitors: {len(COMPETITORS)}")


if __name__ == "__main__":
    build()